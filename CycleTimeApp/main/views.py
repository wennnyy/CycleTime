# ============================================================
# main/views.py
#
# Hanya menangani HTTP request/response cycle.
# Logika bisnis → services.py
# Pure utilities → utils.py
# ============================================================

from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.hashers import make_password
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse, HttpResponseForbidden, JsonResponse
from django.db.models import Avg, Q, Min, Max
from django.db.models import Case, When, Value, IntegerField, Max as MaxId
from django.core.paginator import Paginator
from datetime import date, datetime

from main.models import User, RawTicket, SyncLog, FlagHistory, ErrorTicket
from main.process_groups import enrich_ticket
from main.utils import hitung_cycle_time, get_page_range, format_date_range, role_required
from main.services import (
    hitung_available_records, sync_jira_data,
    get_published_syncs, get_dashboard_filter_options,
    get_published_clean_ticket_queryset,
    paginate_and_enrich_tickets,
    compute_ct_analysis,generate_pdf,
    get_jira_due_date_range,
)

import json
import logging

logger = logging.getLogger(__name__)


# `get_published_syncs` moved to `main.services.get_published_syncs`


# ======================================================
# AUTHENTICATION
# ======================================================
def login_view(request):
    if request.user.is_authenticated:
        return redirect_by_role(request.user)

    storage = messages.get_messages(request)
    list(storage)
    storage.used = True

    error = None
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect_by_role(user)
        else:
            error = "Username atau password salah"

    return render(request, 'main/auth/login.html', {'error': error})


def logout_view(request):
    logout(request)
    return redirect('login')


def redirect_by_role(user):
    if user.role == 'admin':
        return redirect('dashboard_admin')
    elif user.role == 'management':
        return redirect('dashboard_management')
    else:
        return redirect('dashboard_staff')


# ======================================================
# DASHBOARD ADMIN
# ======================================================
# `get_dashboard_filter_options` moved to `main.services.get_dashboard_filter_options`


@login_required
@role_required('admin')
def dashboard_admin(request):
    published = get_published_syncs()

    total_main    = RawTicket.objects.filter(parent_key__isnull=True,  sync_log__in=published).count()
    total_sub     = RawTicket.objects.filter(parent_key__isnull=False, sync_log__in=published).count()
    total_flagged = ErrorTicket.objects.count()
    last_sync     = SyncLog.objects.order_by('-started_at').first()

    context = {
        'username':      request.user.username,
        'role':          request.user.get_role_display(),
        'total_main':    total_main,
        'total_sub':     total_sub,
        'total_flagged': total_flagged,
        'total_users':   User.objects.count(),
        'last_sync':     last_sync,
        **get_dashboard_filter_options(),
    }
    return render(request, 'main/admin/dashboard_admin.html', context)


# ======================================================
# SYNC
# ======================================================
@login_required
@role_required('admin')
def admin_sync(request):

    # untuk membuat tampilan range tgl yang sudah ada di sistem.
    def get_data_range_str():
        dr = RawTicket.objects.filter(due_date__isnull=False).aggregate(
            earliest=Min('due_date'), latest=Max('due_date')
        )
        return format_date_range(dr['earliest'], dr['latest'])
    
    
    # ── POST ─────────────────────────────────────────────────────────────────
    if request.method == 'POST':
        action = request.POST.get('action')

        # ── Sync ─────────────────────────────────────────────────────────────
        if action == 'sync':
            start_date = request.POST.get('start_date')
            end_date   = request.POST.get('end_date')

            sync_log, result = sync_jira_data(request.user, start_date, end_date)

            if result.success:
                if result.error_type == 'empty':
                    messages.warning(request,
                        "⚠️ Tidak ada sub-ticket Completed dalam range tanggal tersebut.")

                elif result.total_processed == 0 and result.total_skipped > 0:
                    # Semua data dalam range ini sudah ada di database
                    sub_text = f"{result.total_skipped_sub} sub-ticket" if result.total_skipped_sub > 0 else "data"
                    messages.warning(request,
                        f"⚠️ Semua ticket dalam range tanggal ini sudah ada di database. "
                        f"({sub_text} dilewati). Tidak ada data baru yang ditambahkan.")

                elif result.total_processed > 0 and result.total_skipped > 0:
                    # Sebagian baru, sebagian sudah ada
                    processed_text = f"{result.total_processed_main} main + {result.total_processed_sub} sub-ticket"
                    messages.success(request,
                        f"✅ Sync berhasil! {processed_text} baru disimpan. "
                        f"⚠️ {result.total_skipped_sub} sub-ticket sudah ada di database (dilewati).")

                else:
                    # Semua data baru, tidak ada yang di-skip
                    processed_text = f"{result.total_processed_main} main + {result.total_processed_sub} sub-ticket"
                    messages.success(request,
                        f"✅ Sync berhasil! {processed_text} baru disimpan.")
            else:
                if result.error_type == 'connection':
                    messages.error(request,
                        f"❌ Sync gagal: Tidak dapat terhubung ke Mock Jira API ({result.error_detail}).")
                elif result.error_type == 'timeout':
                    messages.error(request, "❌ Sync gagal: Request ke Mock Jira API timeout.")
                else:
                    messages.error(request, f"❌ Sync gagal: {result.error_detail}")

            from django.urls import reverse
            url = f"{reverse('admin_sync')}?start_date={start_date}&end_date={end_date}"
            return redirect(url)

        # ── Flag ──────────────────────────────────────────────────────────────
        elif action == 'flag':
            ticket_key = request.POST.get('ticket_key')
            comment    = request.POST.get('comment', '')
            try:
                ticket = RawTicket.objects.get(ticket_key=ticket_key)
                ErrorTicket.objects.get_or_create(
                    ticket=ticket,
                    defaults={
                        'error_message': comment or 'Flagged manually',
                        'flagged_by':    request.user,
                        'raw_payload':   {},
                    }
                )
                FlagHistory.objects.create(
                    ticket=ticket, flagged_by=request.user,
                    action='Flag', comment=comment,
                )
                messages.success(request, f"Ticket {ticket_key} has been flagged.")
            except RawTicket.DoesNotExist:
                messages.error(request, "Ticket not found.")
            return redirect('admin_sync')

        # ── Unflag ────────────────────────────────────────────────────────────
        elif action == 'unflag':
            ticket_key = request.POST.get('ticket_key')
            try:
                ticket = RawTicket.objects.get(ticket_key=ticket_key)
                ErrorTicket.objects.filter(ticket=ticket).delete()
                messages.success(request, f"Flag removed from ticket {ticket_key}.")
            except RawTicket.DoesNotExist:
                messages.error(request, "Ticket not found.")
            return redirect('admin_sync')

        # ── Create Dashboard ──────────────────────────────────────────────────
        elif action == 'create_dashboard':
            flagged_count = ErrorTicket.objects.count()
            if flagged_count > 0:
                messages.error(request,
                    f"❌ Cannot create dashboard: {flagged_count} flagged ticket(s) still exist. "
                    f"Please resolve all flagged tickets in Data Management first."
                )
                return redirect('admin_sync')

            last_sync = SyncLog.objects.filter(status='success').order_by('-started_at').first()
            if not last_sync:
                messages.error(request, "❌ No successful sync data found to publish.")
                return redirect('admin_sync')

            last_sync.status = 'published'
            last_sync.save(update_fields=['status'])
            messages.success(request,
                "✅ Dashboard created successfully! Clean data is now available in Reports.")
            return redirect('admin_sync')

    # ── GET ───────────────────────────────────────────────────────────────────
    today      = date.today()
    start_date = request.GET.get('start_date') or str(today.replace(day=1))
    end_date   = request.GET.get('end_date')   or str(today)
    
    if request.GET.get('records_only') == '1': #get jumlah records tersedia di date range
        total, _ = hitung_available_records(start_date, end_date)
        return JsonResponse({'total': total})

    total, error = hitung_available_records(start_date, end_date)

    query       = request.GET.get('q', '')
    page_number = request.GET.get('page', 1)
    per_page    = int(request.GET.get('per_page', 10))
    if per_page not in [10, 25, 50, 100]:
        per_page = 10

    last_sync     = SyncLog.objects.filter(status='success').order_by('-started_at').first()
    last_any_sync = SyncLog.objects.order_by('-started_at').first()
    flagged_count = ErrorTicket.objects.count()

    sync_data      = []
    paginator      = None
    page_obj       = None
    page_range     = []
    start_index    = 0
    end_index      = 0
    total_filtered = 0
    total_all_raw  = RawTicket.objects.filter(
        sync_log=last_sync, parent_key__isnull=False
    ).count() if last_sync else 0

    if last_sync:
        base_qs = RawTicket.objects.filter(
            parent_key__isnull=False,
            sync_log=last_sync,
        )
        if query:
            base_qs = base_qs.filter(ticket_key__icontains=query)

        latest_ids = (
            base_qs.values('ticket_key')
            .annotate(max_id=MaxId('id'))
            .values_list('max_id', flat=True)
        )
        #menampilkan data error diawal
        qs = RawTicket.objects.filter(id__in=latest_ids).annotate(
            error_priority=Case(
                When(cycle_time__lt=0,        then=Value(0)),
                When(cycle_time__isnull=True, then=Value(1)),
                default=Value(2),
                output_field=IntegerField(),
            )
        ).order_by('error_priority', 'ticket_key')

        flagged_keys   = set(ErrorTicket.objects.values_list('ticket__ticket_key', flat=True))
        total_filtered = qs.count()

        paginator   = Paginator(qs, per_page)
        page_obj    = paginator.get_page(page_number)
        start_index = (page_obj.number - 1) * per_page + 1
        end_index   = min(page_obj.number * per_page, total_filtered)

        for t in page_obj.object_list:
            item = {
                'ticket_key':         t.ticket_key,
                'parent_key':         t.parent_key,
                'platform':           t.platform,
                'package_name':       t.package_name,
                'predefined_process': t.predefined_process,
                'status':             t.status,
                'start_date':         t.start_date,
                'due_date':           t.due_date,
                'cycle_time':         t.cycle_time,
                'has_ct':             t.cycle_time is not None,
                'is_flagged':         t.ticket_key in flagged_keys,
            }
            sync_data.append(enrich_ticket(item))
        page_range = get_page_range(page_obj.number, paginator.num_pages)

    jira_range = get_jira_due_date_range()
    jira_start = jira_range[0] if jira_range else None
    jira_end   = jira_range[1] if jira_range else None
    
    #untuk menampilkan 10 sync log terakhir di bawah tabel
    recent_sync_logs = SyncLog.objects.select_related(
    'admin'
    ).order_by('-started_at')[:30]
    
    context = {
        'start_date':      start_date,
        'end_date':        end_date,
        'total':           total,
        'data_range_str':  get_data_range_str(),
        'jira_start':      jira_start,   
        'jira_end':        jira_end,
        'error':           error,
        'last_sync':       last_sync,
        'sync_data':       sync_data,
        'query':           query,
        'per_page':        per_page,
        'page_obj':        page_obj,
        'paginator':       paginator,
        'page_range':      page_range,
        'start_index':     start_index,
        'end_index':       end_index,
        'total_filtered':  total_filtered,
        'total_all':       total_all_raw,
        'flagged_count':   flagged_count,
        'clean_count':     total_all_raw - flagged_count,
        'has_unpublished': last_sync is not None,
        'has_published':   SyncLog.objects.filter(status='published').exists(),
        'last_skipped':    last_any_sync is not None and last_any_sync.status == 'skipped',
        'last_skipped_count': last_any_sync.total_skipped if last_any_sync and last_any_sync.status == 'skipped' else 0,
        'recent_sync_logs': recent_sync_logs,
    }
    return render(request, 'main/admin/sync.html', context)


# ======================================================
# DATA MANAGEMENT
# ======================================================
@login_required
@role_required('admin')
def admin_data(request):
    query       = request.GET.get('q', '')
    page_number = request.GET.get('page', 1)
    per_page    = int(request.GET.get('per_page', 10))
    if per_page not in [10, 25, 50, 100]:
        per_page = 10

    if request.method == 'POST':
        action     = request.POST.get('action')
        ticket_key = request.POST.get('ticket_key')

        if action == 'resolve':
            try:
                ticket = RawTicket.objects.get(ticket_key=ticket_key)
                ErrorTicket.objects.filter(ticket=ticket).delete()
                last_sync = SyncLog.objects.filter(status='success').order_by('-started_at').first()
                if last_sync:
                    ticket.sync_log = last_sync
                    ticket.save(update_fields=['sync_log'])
                messages.success(request, f"Ticket {ticket_key} has been resolved and returned to sync data.")
            except RawTicket.DoesNotExist:
                messages.error(request, "Ticket not found.")
            return redirect('admin_data')

        elif action == 'edit':
            try:
                ticket = RawTicket.objects.get(ticket_key=ticket_key)
                new_start = request.POST.get('start_date') or None
                new_due   = request.POST.get('due_date')   or None
                new_proc  = request.POST.get('predefined_process') or ticket.predefined_process

                ticket.start_date         = new_start
                ticket.due_date           = new_due
                ticket.predefined_process = new_proc
                ticket.cycle_time         = hitung_cycle_time(new_start, new_due)
                ticket.save()
                messages.success(request, f"Ticket {ticket_key} has been updated.")
            except RawTicket.DoesNotExist:
                messages.error(request, "Ticket not found.")
            return redirect('admin_data')

    flagged_ids = ErrorTicket.objects.values_list('ticket_id', flat=True)
    qs = RawTicket.objects.filter(id__in=flagged_ids).order_by('-id')

    if query:
        qs = qs.filter(ticket_key__icontains=query)

    error_qs       = ErrorTicket.objects.select_related('ticket', 'flagged_by').all()
    error_map      = {e.ticket_id: e.error_message for e in error_qs}
    flagged_by_map = {e.ticket_id: e.flagged_by.username if e.flagged_by else '—' for e in error_qs}

    total_all_raw  = RawTicket.objects.filter(parent_key__isnull=False).count()
    total_flagged  = ErrorTicket.objects.count()
    total_ok       = total_all_raw - total_flagged
    total_filtered = qs.count()

    paginator   = Paginator(qs, per_page)
    page_obj    = paginator.get_page(page_number)
    start_index = (page_obj.number - 1) * per_page + 1
    end_index   = min(page_obj.number * per_page, total_filtered)

    data = []
    for t in page_obj.object_list:
        item = {
            'ticket_key':         t.ticket_key,
            'platform':           t.platform,
            'package_name':       t.package_name,
            'predefined_process': t.predefined_process,
            'status':             t.status,
            'start_date':         t.start_date,
            'due_date':           t.due_date,
            'cycle_time':         t.cycle_time,
            'error_message':      error_map.get(t.id, ''),
            'flagged_by':         flagged_by_map.get(t.id, '—'),
        }
        data.append(enrich_ticket(item))

    context = {
        'data':           data,
        'query':          query,
        'per_page':       per_page,
        'page_obj':       page_obj,
        'paginator':      paginator,
        'page_range':     get_page_range(page_obj.number, paginator.num_pages),
        'start_index':    start_index,
        'end_index':      end_index,
        'total_filtered': total_filtered,
        'total_all':      total_all_raw,
        'total_flagged':  total_flagged,
        'total_ok':       total_ok,
    }
    return render(request, 'main/admin/data.html', context)


# ======================================================
# USER MANAGEMENT
# ======================================================
@login_required
@role_required('admin')
def admin_users(request):
    users = User.objects.all().order_by('username')
    context = {
        'manage_user':      users,
        'total_admin':      users.filter(role='admin').count(),
        'total_staff':      users.filter(role='staff').count(),
        'total_management': users.filter(role='management').count(),
    }
    return render(request, 'main/admin/users.html', context)


@login_required
@role_required('admin')
def add_user(request):
    if request.method == 'POST':
        User.objects.create(
            username=request.POST.get('username'),
            email=request.POST.get('email'),
            role=request.POST.get('role'),
            password=make_password(request.POST.get('password'))
        )
        messages.success(request, "User berhasil ditambahkan")
    return redirect('admin_users')


@login_required
@role_required('admin')
def edit_user(request, user_id):
    user = User.objects.get(id=user_id)
    if request.method == 'POST':
        user.username = request.POST.get('username')
        user.email    = request.POST.get('email')
        user.role     = request.POST.get('role')
        password      = request.POST.get('password')
        if password:
            user.password = make_password(password)
        user.save()
        messages.success(request, "User berhasil diupdate")
    return redirect('admin_users')


@login_required
@role_required('admin')
def delete_user(request, user_id):
    if request.method == 'POST':
        User.objects.get(id=user_id).delete()
        messages.success(request, "User berhasil dihapus")
    return redirect('admin_users')


def _reports_view(request, template_name):
    query       = request.GET.get('q', '')
    date_from   = request.GET.get('date_from', '')
    date_to     = request.GET.get('date_to', '')
    page_number = request.GET.get('page', 1)
    per_page    = int(request.GET.get('per_page', 10))
    if per_page not in [10, 25, 50, 100]:
        per_page = 10

    qs = get_published_clean_ticket_queryset(query, date_from, date_to)

    published = get_published_syncs()
    total_main = RawTicket.objects.filter(
        sync_log__in=published,
        parent_key__isnull=True
    ).count()
    total_sub = RawTicket.objects.filter(
        sync_log__in=published,
        parent_key__isnull=False
    ).count()

    page_data = paginate_and_enrich_tickets(qs, page_number, per_page)
    context = {
        'data':           page_data['data'],
        'query':          query,
        'date_from':      date_from,
        'date_to':        date_to,
        'per_page':       per_page,
        'page_obj':       page_data['page_obj'],
        'paginator':      page_data['paginator'],
        'page_range':     get_page_range(page_data['page_obj'].number, page_data['paginator'].num_pages),
        'start_index':    page_data['start_index'],
        'end_index':      page_data['end_index'],
        'total_filtered': page_data['total_filtered'],
        'total_main':     total_main,
        'total_sub':      total_sub,
    }
    return render(request, template_name, context)


# ======================================================
# REPORTS ADMIN
# ======================================================
@login_required
@role_required('admin')
def admin_reports(request):
    return _reports_view(request, 'main/admin/reports.html')

#endpoint download report excel untuk semua role
@login_required
@role_required('admin', 'staff', 'management')
def download_report_excel(request):
    """Generic Excel report downloader for all roles.
    
    Query params:
    - q: search query
    - date_from: start date filter (YYYY-MM-DD)
    - date_to: end date filter (YYYY-MM-DD)
    """
    query     = request.GET.get('q', '')
    date_from = request.GET.get('date_from', '')
    date_to   = request.GET.get('date_to', '')
    return _generate_report_excel_response(query, date_from, date_to)


def _generate_report_excel_response(query, date_from, date_to):
    import io
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    qs = get_published_clean_ticket_queryset(query, date_from, date_to)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ticket Report"

    COLOR_HEADER_BG = '1E293B'
    COLOR_HEADER_FG = 'FFFFFF'
    COLOR_ROW_ODD   = 'FFFFFF'
    COLOR_ROW_EVEN  = 'F1F5F9'
    COLOR_BORDER    = 'CBD5E1'
    COLOR_ACCENT    = '3B82F6'

    thin_border = Border(
        left=Side(style='thin', color=COLOR_BORDER),
        right=Side(style='thin', color=COLOR_BORDER),
        top=Side(style='thin', color=COLOR_BORDER),
        bottom=Side(style='thin', color=COLOR_BORDER),
    )
    accent_bottom_border = Border(
        left=Side(style='thin', color=COLOR_BORDER),
        right=Side(style='thin', color=COLOR_BORDER),
        top=Side(style='thin', color=COLOR_BORDER),
        bottom=Side(style='medium', color=COLOR_ACCENT),
    )

    ws.merge_cells('A1:L1')
    title_cell = ws['A1']
    title_cell.value     = 'Clean Ticket Report'
    title_cell.font      = Font(name='Calibri', bold=True, size=14, color=COLOR_HEADER_BG)
    title_cell.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 24

    ws.merge_cells('A2:L2')
    info_parts = []
    if query:     info_parts.append(f"Search: {query}")
    if date_from: info_parts.append(f"From: {date_from}")
    if date_to:   info_parts.append(f"To: {date_to}")
    info_parts.append(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    info_parts.append(f"Total records: {qs.count()}")
    info_cell = ws['A2']
    info_cell.value     = '  |  '.join(info_parts)
    info_cell.font      = Font(name='Calibri', size=9, color='64748B')
    info_cell.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 6

    HEADERS = [
        ('No', 5), ('Ticket ID', 14), ('Parent', 14), ('Package', 14), ('Platform', 14),
        ('Process', 18), ('Stage', 14), ('Area', 16), ('Start Date', 13),
        ('Due Date', 13), ('CT (days)', 10), ('Status', 13),
    ]
    header_fill  = PatternFill('solid', fgColor=COLOR_HEADER_BG)
    header_font  = Font(name='Calibri', bold=True, size=9, color=COLOR_HEADER_FG)
    header_align = Alignment(horizontal='center', vertical='center')

    for col_idx, (label, width) in enumerate(HEADERS, start=1):
        cell = ws.cell(row=4, column=col_idx, value=label)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = header_align
        cell.border    = accent_bottom_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[4].height = 20

    center_align = Alignment(horizontal='center', vertical='center')
    left_align   = Alignment(horizontal='left',   vertical='center')
    ct_font      = Font(name='Calibri', size=9, bold=True, color='1D4ED8')
    base_font    = Font(name='Calibri', size=9)

    for i, t in enumerate(qs, start=1):
        row_num = i + 4
        item = {
            'ticket_key':         t.ticket_key,
            'parent_key':         t.parent_key,
            'platform':           t.platform,
            'package_name':       t.package_name,
            'predefined_process': t.predefined_process,
            'status':             t.status,
            'start_date':         t.start_date,
            'due_date':           t.due_date,
            'cycle_time':         t.cycle_time,
        }
        e      = enrich_ticket(item)
        ct_val = round(e['cycle_time'], 1) if e.get('cycle_time') else None
        row_fill = PatternFill('solid', fgColor=COLOR_ROW_ODD if i % 2 else COLOR_ROW_EVEN)

        values = [
            i, e.get('ticket_key',''), e.get('parent_key') or '',
            e.get('package_name') or '', e.get('platform_group') or '',
            e.get('predefined_process') or '', e.get('process_stage') or '',
            e.get('process_area') or '', str(e.get('start_date') or ''),
            str(e.get('due_date') or ''), ct_val, e.get('status',''),
        ]
        aligns = [
            center_align, left_align, left_align, left_align, left_align,
            left_align, left_align, left_align,
            center_align, center_align, center_align, center_align,
        ]
        for col_idx, (val, aln) in enumerate(zip(values, aligns), start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.fill      = row_fill
            cell.border    = thin_border
            cell.alignment = aln
            cell.font      = ct_font if col_idx == 11 and val is not None else base_font
        ws.row_dimensions[row_num].height = 16

    ws.freeze_panes = 'A5'
    ws.auto_filter.ref = f"A4:L{ws.max_row}"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"report-{datetime.now().strftime('%Y%m%d-%H%M%S')}.xlsx"
    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ======================================================
# DASHBOARD STAFF
# ======================================================
@login_required
@role_required('staff')
def dashboard_staff(request):
    published     = get_published_syncs()
    total_ticket  = RawTicket.objects.filter(parent_key__isnull=False, sync_log__in=published).count()
    avg_ct        = RawTicket.objects.filter(cycle_time__isnull=False, sync_log__in=published).aggregate(avg=Avg('cycle_time'))['avg']
    avg_ct        = round(avg_ct, 1) if avg_ct else 0
    total_flagged = ErrorTicket.objects.count()

    context = {
        'username':      request.user.username,
        'role':          request.user.get_role_display(),
        'total_ticket':  total_ticket,
        'avg_ct':        avg_ct,
        'total_flagged': total_flagged,
        **get_dashboard_filter_options(),
    }
    return render(request, 'main/staff/dashboard_staff.html', context)


# ======================================================
# VIEW DATA STAFF
# ======================================================
@login_required
@role_required('staff')
def staff_view_data(request):
    query       = request.GET.get('q', '')
    page_number = request.GET.get('page', 1)
    per_page    = int(request.GET.get('per_page', 10))
    if per_page not in [10, 25, 50, 100]:
        per_page = 10

    if request.method == 'POST':
        action     = request.POST.get('action')
        ticket_key = request.POST.get('ticket_key')
        comment    = request.POST.get('comment', '')

        if action == 'flag':
            try:
                ticket = RawTicket.objects.get(ticket_key=ticket_key)
                ErrorTicket.objects.get_or_create(
                    ticket=ticket,
                    defaults={
                        'error_message': comment or 'Flagged by staff',
                        'flagged_by':    request.user,
                        'raw_payload':   {},
                    }
                )
                FlagHistory.objects.create(
                    ticket=ticket, flagged_by=request.user,
                    action='Flag', comment=comment,
                )
                messages.success(request, f"Ticket {ticket_key} has been flagged.")
            except RawTicket.DoesNotExist:
                messages.error(request, "Ticket not found.")
        return redirect('staff_view_data')

    last_sync      = SyncLog.objects.filter(status='success').order_by('-started_at').first()
    sync_data      = []
    paginator      = None
    page_obj       = None
    page_range     = []
    start_index    = 0
    end_index      = 0
    total_filtered = 0
    total_all_raw  = RawTicket.objects.filter(parent_key__isnull=False).count()

    if last_sync:
        #untuk membuat minus muncul di awal
        qs = RawTicket.objects.filter(
            parent_key__isnull=False,
            sync_log=last_sync,
        ).annotate(
            error_priority=Case(
                When(cycle_time__lt=0,        then=Value(0)),
                When(cycle_time__isnull=True, then=Value(1)),
                default=Value(2),
                output_field=IntegerField(),
            )
        ).order_by('error_priority', 'ticket_key')

        if query:
            qs = qs.filter(
                Q(ticket_key__icontains=query) |
                Q(predefined_process__icontains=query)
            )

        flagged_keys   = set(ErrorTicket.objects.values_list('ticket__ticket_key', flat=True))
        total_filtered = qs.count()

        paginator   = Paginator(qs, per_page)
        page_obj    = paginator.get_page(page_number)
        start_index = (page_obj.number - 1) * per_page + 1
        end_index   = min(page_obj.number * per_page, total_filtered)

        for t in page_obj.object_list:
            item = {
                'ticket_key':         t.ticket_key,
                'parent_key':         t.parent_key,
                'platform':           t.platform,
                'package_name':       t.package_name,
                'predefined_process': t.predefined_process,
                'status':             t.status,
                'start_date':         t.start_date,
                'due_date':           t.due_date,
                'cycle_time':         t.cycle_time,
                'has_ct':             t.cycle_time is not None,
                'is_flagged':         t.ticket_key in flagged_keys,
            }
            sync_data.append(enrich_ticket(item))
        page_range = get_page_range(page_obj.number, paginator.num_pages)

   # ── Data range string ─────────────────────────────────────
    dr = RawTicket.objects.filter(due_date__isnull=False).aggregate(
        earliest=Min('due_date'), latest=Max('due_date')
    )
    data_range_str = format_date_range(dr['earliest'], dr['latest'])
    
    total_all_sync  = RawTicket.objects.filter(
        parent_key__isnull=False, sync_log=last_sync
    ).count() if last_sync else 0
    flagged_count   = ErrorTicket.objects.count()
    clean_count     = total_all_sync - flagged_count
    context = {
        'sync_data':      sync_data,
        'query':          query,
        'per_page':       per_page,
        'page_obj':       page_obj,
        'paginator':      paginator,
        'page_range':     page_range,
        'start_index':    start_index,
        'end_index':      end_index,
        'total_filtered': total_filtered,
        'total_all':      total_all_raw,
        'last_sync':      last_sync,
        'data_range_str': data_range_str,
        'clean_count':    clean_count,
        'flagged_count':  flagged_count,
    }
    return render(request, 'main/staff/view_data.html', context)


# ======================================================
# REPORTS STAFF
# ======================================================
@login_required
@role_required('staff')
def staff_reports(request):
    return _reports_view(request, 'main/staff/reports.html')


# ======================================================
# DASHBOARD MANAGEMENT
# ======================================================
@login_required
@role_required('management')
def dashboard_management(request):
    published    = get_published_syncs()
    total_ticket = RawTicket.objects.filter(parent_key__isnull=False, sync_log__in=published).count()
    avg_ct       = RawTicket.objects.filter(cycle_time__isnull=False, sync_log__in=published).aggregate(avg=Avg('cycle_time'))['avg']
    avg_ct       = round(avg_ct, 1) if avg_ct else 0

    context = {
        'username':     request.user.username,
        'role':         request.user.get_role_display(),
        'total_ticket': total_ticket,
        'avg_ct':       avg_ct,
        **get_dashboard_filter_options(),
    }
    return render(request, 'main/management/dashboard_management.html', context)


# ======================================================
# REPORTS MANAGEMENT
# ======================================================
@login_required
@role_required('management')
def management_reports(request):
    return _reports_view(request, 'main/management/reports.html')



# ======================================================
# CYCLE TIME ANALYSIS DASHBOARD
# ======================================================
@login_required
def ct_analysis_data(request):
    f_year      = request.GET.get('year', '').strip()
    f_platforms = request.GET.getlist('platform')
    f_stages    = request.GET.getlist('stage')
    f_areas     = request.GET.getlist('area')
    f_processes = request.GET.getlist('process')

    data = compute_ct_analysis(f_year, f_platforms, f_stages, f_areas, f_processes)
    return JsonResponse(data)


@login_required
def download_dashboard_pdf(request):
    if request.user.role not in ('admin', 'staff', 'management'):
        return HttpResponseForbidden("Akses ditolak.")
    if request.method == 'POST':
        # Frontend kirim data hasil compute sebelumnya sebagai JSON
        payload  = json.loads(request.body)
        data     = payload.get('data', {})
        filters  = payload.get('filters', {})
    else:
        # Fallback: GET tetap bisa dipakai
        filters = {
            'year':      request.GET.get('year', '').strip(),
            'platforms': request.GET.getlist('platform'),
            'stages':    request.GET.getlist('stage'),
            'areas':     request.GET.getlist('area'),
            'processes': request.GET.getlist('process'),
        }
        data = compute_ct_analysis(
            filters['year'], filters['platforms'],
            filters['stages'], filters['areas'], filters['processes'],
        )

    pdf_buffer = generate_pdf(data, filters)
    filename   = f"ct-analysis-{datetime.now().strftime('%Y%m%d-%H%M')}.pdf"
    response   = HttpResponse(pdf_buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response