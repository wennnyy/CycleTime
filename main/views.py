# ============================================================
# main/views.py
#
# Hanya menangani HTTP request/response cycle.
# Logika bisnis → services.py
# Pure utilities → utils.py
# ============================================================

from functools import wraps
from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.hashers import make_password
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse, HttpResponseForbidden, JsonResponse
from django.utils import timezone
from django.db.models import Avg, Count, Q, Min, Max
from django.db.models import Case, When, Value, IntegerField, Max as MaxId
from django.core.paginator import Paginator
from django.conf import settings
from datetime import date, datetime

from requests import request

from main.models import User, RawTicket, SyncLog, FlagHistory, ErrorTicket
from main.process_groups import enrich_ticket, get_platform
from main.utils import hitung_cycle_time, get_page_range
from main.services import hitung_available_records, sync_jira_data

import logging

logger = logging.getLogger(__name__)


# ======================================================
# HELPER: ROLE DECORATOR
# ======================================================
def role_required(role_name):
    def decorator(view_func):
        @wraps(view_func)
        def wrapper(request, *args, **kwargs):
            if not request.user.is_authenticated:
                return redirect('login')
            if request.user.role != role_name:
                return HttpResponseForbidden("⛔ Anda tidak memiliki akses ke halaman ini")
            return view_func(request, *args, **kwargs)
        return wrapper
    return decorator


def get_published_syncs():
    """Hanya data yang sudah di-publish via tombol Create Dashboard."""
    return SyncLog.objects.filter(status='published').values_list('id', flat=True)


# ======================================================
# AUTHENTICATION
# ======================================================
def login_view(request):
    if request.user.is_authenticated:
        return redirect_by_role(request.user)

    storage = messages.get_messages(request)
    list(storage)
    storage.used = True

    error = None
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect_by_role(user)
        else:
            error = "Username atau password salah"

    return render(request, 'main/auth/login.html', {'error': error})


def logout_view(request):
    logout(request)
    return redirect('login')


def redirect_by_role(user):
    if user.role == 'admin':
        return redirect('dashboard_admin')
    elif user.role == 'management':
        return redirect('dashboard_management')
    else:
        return redirect('dashboard_staff')


# ======================================================
# DASHBOARD ADMIN
# ======================================================
def get_dashboard_filter_options():
    """Ambil semua nilai unik untuk filter dropdown di dashboard."""
    from main.process_groups import PROCESS_GROUP, PACKAGE_PLATFORM
    published = get_published_syncs()

    years_qs = RawTicket.objects.filter(
        parent_key__isnull=False,
        start_date__isnull=False,
        cycle_time__isnull=False,
        sync_log__in=published,
    ).dates('start_date', 'year')

    return {
        'all_years':     sorted(set(str(d.year) for d in years_qs)),
        'all_platforms': sorted(set(PACKAGE_PLATFORM.values())),
        'all_stages':    sorted(set(pg[0] for pg in PROCESS_GROUP.values())),
        'all_areas':     sorted(set(pg[1] for pg in PROCESS_GROUP.values())),
        'all_processes': sorted(set(PROCESS_GROUP.keys())),
    }


@login_required
@role_required('admin')
def dashboard_admin(request):
    published = get_published_syncs()

    total_main    = RawTicket.objects.filter(parent_key__isnull=True,  sync_log__in=published).count()
    total_sub     = RawTicket.objects.filter(parent_key__isnull=False, sync_log__in=published).count()
    total_flagged = ErrorTicket.objects.count()
    last_sync     = SyncLog.objects.order_by('-started_at').first()

    context = {
        'username':      request.user.username,
        'role':          request.user.get_role_display(),
        'total_main':    total_main,
        'total_sub':     total_sub,
        'total_flagged': total_flagged,
        'total_users':   User.objects.count(),
        'last_sync':     last_sync,
        **get_dashboard_filter_options(),
    }
    return render(request, 'main/admin/dashboard_admin.html', context)


# ======================================================
# SYNC
# ======================================================
@login_required
@role_required('admin')
def admin_sync(request):

    BULAN_ID = {
        1:'Januari', 2:'Februari', 3:'Maret',    4:'April',
        5:'Mei',     6:'Juni',     7:'Juli',      8:'Agustus',
        9:'September',10:'Oktober',11:'November', 12:'Desember',
    }
    # untuk membuat tampilan range tgl yang sudah ada di sistem.
    def get_data_range_str():
        dr = RawTicket.objects.filter(start_date__isnull=False).aggregate(
            earliest=Min('start_date'), latest=Max('start_date')
        )
        e, l = dr['earliest'], dr['latest']
        if e and l:
            # SEBELUM: hanya bulan dan tahun
            # return f"{BULAN_ID[e.month]} {e.year} — {BULAN_ID[l.month]} {l.year}"
            # SESUDAH: tanggal lengkap
            return f"{e.day} {BULAN_ID[e.month]} {e.year} — {l.day} {BULAN_ID[l.month]} {l.year}"
        return None

    # ── POST ─────────────────────────────────────────────────────────────────
    if request.method == 'POST':
        action = request.POST.get('action')

        # ── Sync ─────────────────────────────────────────────────────────────
        if action == 'sync':
            start_date = request.POST.get('start_date')
            end_date   = request.POST.get('end_date')

            sync_log, result = sync_jira_data(request.user, start_date, end_date)

            if result.success:
                if result.error_type == 'empty':
                    messages.warning(request,
                        "⚠️ Tidak ada sub-ticket Completed dalam range tanggal tersebut.")
                elif result.total_skipped > 0 and result.total_processed == 0:
                    messages.warning(request,
                        f"⚠️ Semua {result.total_skipped} ticket di range ini sudah ada di database. "
                        f"Tidak ada data baru yang ditambahkan.")
                elif result.total_skipped > 0:
                    messages.success(request,
                        f"✅ Sync berhasil! {result.total_processed} ticket baru disimpan. "
                        f"{result.total_skipped} ticket sudah ada di database (dilewati).")
                else:
                    messages.success(request,
                        f"✅ Sync berhasil! {result.total_processed} ticket baru disimpan.")
            else:
                if result.error_type == 'connection':
                    messages.error(request,
                        f"❌ Sync gagal: Tidak dapat terhubung ke Mock Jira API ({result.error_detail}).")
                elif result.error_type == 'timeout':
                    messages.error(request, "❌ Sync gagal: Request ke Mock Jira API timeout.")
                else:
                    messages.error(request, f"❌ Sync gagal: {result.error_detail}")

            from django.urls import reverse
            url = f"{reverse('admin_sync')}?start_date={start_date}&end_date={end_date}"
            return redirect(url)

        # ── Flag ──────────────────────────────────────────────────────────────
        elif action == 'flag':
            ticket_key = request.POST.get('ticket_key')
            comment    = request.POST.get('comment', '')
            try:
                ticket = RawTicket.objects.get(ticket_key=ticket_key)
                ErrorTicket.objects.get_or_create(
                    ticket=ticket,
                    defaults={
                        'error_message': comment or 'Flagged manually',
                        'flagged_by':    request.user,
                        'raw_payload':   {},
                    }
                )
                FlagHistory.objects.create(
                    ticket=ticket, flagged_by=request.user,
                    action='Flag', comment=comment,
                )
                messages.success(request, f"Ticket {ticket_key} has been flagged.")
            except RawTicket.DoesNotExist:
                messages.error(request, "Ticket not found.")
            return redirect('admin_sync')

        # ── Unflag ────────────────────────────────────────────────────────────
        elif action == 'unflag':
            ticket_key = request.POST.get('ticket_key')
            try:
                ticket = RawTicket.objects.get(ticket_key=ticket_key)
                ErrorTicket.objects.filter(ticket=ticket).delete()
                messages.success(request, f"Flag removed from ticket {ticket_key}.")
            except RawTicket.DoesNotExist:
                messages.error(request, "Ticket not found.")
            return redirect('admin_sync')

        # ── Create Dashboard ──────────────────────────────────────────────────
        elif action == 'create_dashboard':
            flagged_count = ErrorTicket.objects.count()
            if flagged_count > 0:
                messages.error(request,
                    f"❌ Cannot create dashboard: {flagged_count} flagged ticket(s) still exist. "
                    f"Please resolve all flagged tickets in Data Management first."
                )
                return redirect('admin_sync')

            last_sync = SyncLog.objects.filter(status='success').order_by('-started_at').first()
            if not last_sync:
                messages.error(request, "❌ No successful sync data found to publish.")
                return redirect('admin_sync')

            last_sync.status = 'published'
            last_sync.save(update_fields=['status'])
            messages.success(request,
                "✅ Dashboard created successfully! Clean data is now available in Reports.")
            return redirect('admin_sync')

    # ── GET ───────────────────────────────────────────────────────────────────
    today      = date.today()
    start_date = request.GET.get('start_date') or str(today.replace(day=1))
    end_date   = request.GET.get('end_date')   or str(today)
    
    if request.GET.get('records_only') == '1':
        total, _ = hitung_available_records(start_date, end_date)
        return JsonResponse({'total': total})

    total, error = hitung_available_records(start_date, end_date)

    query       = request.GET.get('q', '')
    page_number = request.GET.get('page', 1)
    per_page    = int(request.GET.get('per_page', 10))
    if per_page not in [10, 25, 50, 100]:
        per_page = 10

    last_sync     = SyncLog.objects.filter(status='success').order_by('-started_at').first()
    flagged_count = ErrorTicket.objects.count()

    sync_data      = []
    paginator      = None
    page_obj       = None
    page_range     = []
    start_index    = 0
    end_index      = 0
    total_filtered = 0
    total_all_raw  = RawTicket.objects.filter(
        sync_log=last_sync, parent_key__isnull=False
    ).count() if last_sync else 0

    if last_sync:
        base_qs = RawTicket.objects.filter(
            parent_key__isnull=False,
            sync_log=last_sync,
        )
        if query:
            base_qs = base_qs.filter(ticket_key__icontains=query)

        latest_ids = (
            base_qs.values('ticket_key')
            .annotate(max_id=MaxId('id'))
            .values_list('max_id', flat=True)
        )

        qs = RawTicket.objects.filter(id__in=latest_ids).annotate(
            error_priority=Case(
                When(cycle_time__lt=0,        then=Value(0)),
                When(cycle_time__isnull=True, then=Value(1)),
                default=Value(2),
                output_field=IntegerField(),
            )
        ).order_by('error_priority', 'ticket_key')

        flagged_keys   = set(ErrorTicket.objects.values_list('ticket__ticket_key', flat=True))
        total_filtered = qs.count()

        paginator   = Paginator(qs, per_page)
        page_obj    = paginator.get_page(page_number)
        start_index = (page_obj.number - 1) * per_page + 1
        end_index   = min(page_obj.number * per_page, total_filtered)

        for t in page_obj.object_list:
            item = {
                'ticket_key':         t.ticket_key,
                'parent_key':         t.parent_key,
                'platform':           t.platform,
                'package_name':       t.package_name,
                'predefined_process': t.predefined_process,
                'status':             t.status,
                'start_date':         t.start_date,
                'due_date':           t.due_date,
                'cycle_time':         t.cycle_time,
                'has_ct':             t.cycle_time is not None,
                'is_flagged':         t.ticket_key in flagged_keys,
            }
            sync_data.append(enrich_ticket(item))
        page_range = get_page_range(page_obj.number, paginator.num_pages)

    context = {
        'start_date':      start_date,
        'end_date':        end_date,
        'total':           total,
        'data_range_str':  get_data_range_str(),
        'error':           error,
        'last_sync':       last_sync,
        'sync_data':       sync_data,
        'query':           query,
        'per_page':        per_page,
        'page_obj':        page_obj,
        'paginator':       paginator,
        'page_range':      page_range,
        'start_index':     start_index,
        'end_index':       end_index,
        'total_filtered':  total_filtered,
        'total_all':       total_all_raw,
        'flagged_count':   flagged_count,
        'clean_count':     total_all_raw - flagged_count,
        'has_unpublished': last_sync is not None,
        'has_published':   SyncLog.objects.filter(status='published').exists(),
    }
    return render(request, 'main/admin/sync.html', context)


# ======================================================
# DATA MANAGEMENT
# ======================================================
@login_required
@role_required('admin')
def admin_data(request):
    query       = request.GET.get('q', '')
    page_number = request.GET.get('page', 1)
    per_page    = int(request.GET.get('per_page', 10))
    if per_page not in [10, 25, 50, 100]:
        per_page = 10

    if request.method == 'POST':
        action     = request.POST.get('action')
        ticket_key = request.POST.get('ticket_key')

        if action == 'resolve':
            try:
                ticket = RawTicket.objects.get(ticket_key=ticket_key)
                ErrorTicket.objects.filter(ticket=ticket).delete()
                last_sync = SyncLog.objects.filter(status='success').order_by('-started_at').first()
                if last_sync:
                    ticket.sync_log = last_sync
                    ticket.save(update_fields=['sync_log'])
                messages.success(request, f"Ticket {ticket_key} has been resolved and returned to sync data.")
            except RawTicket.DoesNotExist:
                messages.error(request, "Ticket not found.")
            return redirect('admin_data')

        elif action == 'edit':
            try:
                ticket = RawTicket.objects.get(ticket_key=ticket_key)
                new_start = request.POST.get('start_date') or None
                new_due   = request.POST.get('due_date')   or None
                new_proc  = request.POST.get('predefined_process') or ticket.predefined_process

                ticket.start_date         = new_start
                ticket.due_date           = new_due
                ticket.predefined_process = new_proc
                ticket.cycle_time         = hitung_cycle_time(new_start, new_due)
                ticket.save()
                messages.success(request, f"Ticket {ticket_key} has been updated.")
            except RawTicket.DoesNotExist:
                messages.error(request, "Ticket not found.")
            return redirect('admin_data')

    flagged_ids = ErrorTicket.objects.values_list('ticket_id', flat=True)
    qs = RawTicket.objects.filter(id__in=flagged_ids).order_by('-id')

    if query:
        qs = qs.filter(ticket_key__icontains=query)

    error_qs       = ErrorTicket.objects.select_related('ticket', 'flagged_by').all()
    error_map      = {e.ticket_id: e.error_message for e in error_qs}
    flagged_by_map = {e.ticket_id: e.flagged_by.username if e.flagged_by else '—' for e in error_qs}

    total_all_raw  = RawTicket.objects.filter(parent_key__isnull=False).count()
    total_flagged  = ErrorTicket.objects.count()
    total_ok       = total_all_raw - total_flagged
    total_filtered = qs.count()

    paginator   = Paginator(qs, per_page)
    page_obj    = paginator.get_page(page_number)
    start_index = (page_obj.number - 1) * per_page + 1
    end_index   = min(page_obj.number * per_page, total_filtered)

    data = []
    for t in page_obj.object_list:
        item = {
            'ticket_key':         t.ticket_key,
            'platform':           t.platform,
            'package_name':       t.package_name,
            'predefined_process': t.predefined_process,
            'status':             t.status,
            'start_date':         t.start_date,
            'due_date':           t.due_date,
            'cycle_time':         t.cycle_time,
            'error_message':      error_map.get(t.id, ''),
            'flagged_by':         flagged_by_map.get(t.id, '—'),
        }
        data.append(enrich_ticket(item))

    context = {
        'data':           data,
        'query':          query,
        'per_page':       per_page,
        'page_obj':       page_obj,
        'paginator':      paginator,
        'page_range':     get_page_range(page_obj.number, paginator.num_pages),
        'start_index':    start_index,
        'end_index':      end_index,
        'total_filtered': total_filtered,
        'total_all':      total_all_raw,
        'total_flagged':  total_flagged,
        'total_ok':       total_ok,
    }
    return render(request, 'main/admin/data.html', context)


# ======================================================
# USER MANAGEMENT
# ======================================================
@login_required
@role_required('admin')
def admin_users(request):
    users = User.objects.all().order_by('username')
    context = {
        'manage_user':      users,
        'total_admin':      users.filter(role='admin').count(),
        'total_staff':      users.filter(role='staff').count(),
        'total_management': users.filter(role='management').count(),
    }
    return render(request, 'main/admin/users.html', context)


@login_required
@role_required('admin')
def add_user(request):
    if request.method == 'POST':
        User.objects.create(
            username=request.POST.get('username'),
            email=request.POST.get('email'),
            role=request.POST.get('role'),
            password=make_password(request.POST.get('password'))
        )
        messages.success(request, "User berhasil ditambahkan")
    return redirect('admin_users')


@login_required
@role_required('admin')
def edit_user(request, user_id):
    user = User.objects.get(id=user_id)
    if request.method == 'POST':
        user.username = request.POST.get('username')
        user.email    = request.POST.get('email')
        user.role     = request.POST.get('role')
        password      = request.POST.get('password')
        if password:
            user.password = make_password(password)
        user.save()
        messages.success(request, "User berhasil diupdate")
    return redirect('admin_users')


@login_required
@role_required('admin')
def delete_user(request, user_id):
    User.objects.get(id=user_id).delete()
    messages.success(request, "User berhasil dihapus")
    return redirect('admin_users')


# ======================================================
# REPORTS ADMIN
# ======================================================
@login_required
@role_required('admin')
def admin_reports(request):
    query       = request.GET.get('q', '')
    date_from   = request.GET.get('date_from', '')
    date_to     = request.GET.get('date_to', '')
    page_number = request.GET.get('page', 1)
    per_page    = int(request.GET.get('per_page', 10))
    if per_page not in [10, 25, 50, 100]:
        per_page = 10

    published_syncs = SyncLog.objects.filter(status='published').values_list('id', flat=True)
    qs = RawTicket.objects.filter(
        sync_log__in=published_syncs, parent_key__isnull=False
    ).order_by('ticket_key')

    if query:
        qs = qs.filter(
            Q(ticket_key__icontains=query) |
            Q(platform__icontains=query)   |
            Q(predefined_process__icontains=query)
        )
    if date_from:
        qs = qs.filter(start_date__gte=date_from)
    if date_to:
        qs = qs.filter(start_date__lte=date_to)

    seen, unique_ids = set(), []
    for t in qs.values('id', 'ticket_key'):
        if t['ticket_key'] not in seen:
            seen.add(t['ticket_key'])
            unique_ids.append(t['id'])
    qs = RawTicket.objects.filter(id__in=unique_ids).order_by('ticket_key')

    total_main     = RawTicket.objects.filter(sync_log__in=published_syncs, parent_key__isnull=True).count()
    total_sub      = RawTicket.objects.filter(sync_log__in=published_syncs, parent_key__isnull=False).count()
    total_filtered = qs.count()

    paginator   = Paginator(qs, per_page)
    page_obj    = paginator.get_page(page_number)
    start_index = (page_obj.number - 1) * per_page + 1
    end_index   = min(page_obj.number * per_page, total_filtered)

    data = []
    for t in page_obj.object_list:
        item = {
            'ticket_key':         t.ticket_key,
            'parent_key':         t.parent_key,
            'platform':           t.platform,
            'package_name':       t.package_name,
            'predefined_process': t.predefined_process,
            'status':             t.status,
            'start_date':         t.start_date,
            'due_date':           t.due_date,
            'cycle_time':         t.cycle_time,
            'has_ct':             t.cycle_time is not None,
        }
        data.append(enrich_ticket(item))

    context = {
        'data':           data,
        'query':          query,
        'date_from':      date_from,
        'date_to':        date_to,
        'per_page':       per_page,
        'page_obj':       page_obj,
        'paginator':      paginator,
        'page_range':     get_page_range(page_obj.number, paginator.num_pages),
        'start_index':    start_index,
        'end_index':      end_index,
        'total_filtered': total_filtered,
        'total_main':     total_main,
        'total_sub':      total_sub,
    }
    return render(request, 'main/admin/reports.html', context)


@login_required
@role_required('admin')
def download_report(request):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    import io

    query     = request.GET.get('q', '')
    date_from = request.GET.get('date_from', '')
    date_to   = request.GET.get('date_to', '')

    published_syncs = SyncLog.objects.filter(status='published').values_list('id', flat=True)
    qs = RawTicket.objects.filter(
        sync_log__in=published_syncs, parent_key__isnull=False
    ).order_by('ticket_key')

    if query:
        qs = qs.filter(
            Q(ticket_key__icontains=query) |
            Q(platform__icontains=query)   |
            Q(predefined_process__icontains=query)
        )
    if date_from:
        qs = qs.filter(start_date__gte=date_from)
    if date_to:
        qs = qs.filter(start_date__lte=date_to)

    seen, unique_ids = set(), []
    for t in qs.values('id', 'ticket_key'):
        if t['ticket_key'] not in seen:
            seen.add(t['ticket_key'])
            unique_ids.append(t['id'])
    qs = RawTicket.objects.filter(id__in=unique_ids).order_by('ticket_key')

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(A4),
        leftMargin=1*cm, rightMargin=1*cm,
        topMargin=1.5*cm, bottomMargin=1.5*cm,
    )
    styles   = getSampleStyleSheet()
    elements = []

    title_style = ParagraphStyle(
        'ReportTitle', parent=styles['Heading1'],
        fontSize=14, spaceAfter=4, textColor=colors.HexColor('#1e293b'),
    )
    elements.append(Paragraph("Clean Ticket Report", title_style))

    sub_parts = []
    if query:     sub_parts.append(f"Search: {query}")
    if date_from: sub_parts.append(f"From: {date_from}")
    if date_to:   sub_parts.append(f"To: {date_to}")
    sub_parts.append(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    sub_parts.append(f"Total records: {qs.count()}")
    elements.append(Paragraph("  |  ".join(sub_parts), styles['Normal']))
    elements.append(Spacer(1, 0.5*cm))

    headers = [
        'No', 'Ticket ID', 'Parent', 'Package', 'Platform',
        'Process', 'Stage', 'Area', 'Start Date', 'Due Date', 'CT (days)', 'Status',
    ]
    table_data = [headers]

    for i, t in enumerate(qs, start=1):
        item = {
            'ticket_key':         t.ticket_key,
            'parent_key':         t.parent_key,
            'platform':           t.platform,
            'package_name':       t.package_name,
            'predefined_process': t.predefined_process,
            'status':             t.status,
            'start_date':         t.start_date,
            'due_date':           t.due_date,
            'cycle_time':         t.cycle_time,
        }
        enriched = enrich_ticket(item)
        ct_val   = f"{enriched['cycle_time']:.1f}" if enriched.get('cycle_time') else '—'
        table_data.append([
            str(i),
            enriched.get('ticket_key', '—'),
            enriched.get('parent_key') or '—',
            enriched.get('platform') or '—',
            enriched.get('platform_group') or '—',
            enriched.get('predefined_process') or '—',
            enriched.get('process_stage') or '—',
            enriched.get('process_area') or '—',
            str(enriched.get('start_date') or '—'),
            str(enriched.get('due_date') or '—'),
            ct_val,
            enriched.get('status', '—'),
        ])

    col_widths = [
        0.8*cm, 2.8*cm, 2.8*cm, 2.5*cm, 2.5*cm,
        3.2*cm, 2.5*cm, 2.5*cm, 2.2*cm, 2.2*cm, 1.8*cm, 2.2*cm,
    ]
    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND',    (0, 0), (-1, 0),  colors.HexColor('#1e293b')),
        ('TEXTCOLOR',     (0, 0), (-1, 0),  colors.white),
        ('FONTNAME',      (0, 0), (-1, 0),  'Helvetica-Bold'),
        ('FONTSIZE',      (0, 0), (-1, 0),  7),
        ('FONTSIZE',      (0, 1), (-1, -1), 6.5),
        ('FONTNAME',      (0, 1), (-1, -1), 'Helvetica'),
        ('ALIGN',         (0, 0), (-1, -1), 'CENTER'),
        ('ALIGN',         (1, 1), (2, -1),  'LEFT'),
        ('ALIGN',         (5, 1), (7, -1),  'LEFT'),
        ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS',(0, 1), (-1, -1), [colors.white, colors.HexColor('#f1f5f9')]),
        ('GRID',          (0, 0), (-1, -1), 0.4, colors.HexColor('#cbd5e1')),
        ('LINEBELOW',     (0, 0), (-1, 0),  1.2, colors.HexColor('#3b82f6')),
        ('TOPPADDING',    (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('LEFTPADDING',   (0, 0), (-1, -1), 4),
        ('RIGHTPADDING',  (0, 0), (-1, -1), 4),
    ]))
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)

    filename = f"report-{datetime.now().strftime('%Y%m%d-%H%M%S')}.pdf"
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
@role_required('admin')
def download_report_excel(request):
    import io
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    query     = request.GET.get('q', '')
    date_from = request.GET.get('date_from', '')
    date_to   = request.GET.get('date_to', '')

    published_syncs = SyncLog.objects.filter(status='published').values_list('id', flat=True)
    qs = RawTicket.objects.filter(
        sync_log__in=published_syncs, parent_key__isnull=False
    ).order_by('ticket_key')

    if query:
        qs = qs.filter(
            Q(ticket_key__icontains=query) |
            Q(platform__icontains=query)   |
            Q(predefined_process__icontains=query)
        )
    if date_from:
        qs = qs.filter(start_date__gte=date_from)
    if date_to:
        qs = qs.filter(start_date__lte=date_to)

    seen, unique_ids = set(), []
    for t in qs.values('id', 'ticket_key'):
        if t['ticket_key'] not in seen:
            seen.add(t['ticket_key'])
            unique_ids.append(t['id'])
    qs = RawTicket.objects.filter(id__in=unique_ids).order_by('ticket_key')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ticket Report"

    COLOR_HEADER_BG = '1E293B'
    COLOR_HEADER_FG = 'FFFFFF'
    COLOR_ROW_ODD   = 'FFFFFF'
    COLOR_ROW_EVEN  = 'F1F5F9'
    COLOR_BORDER    = 'CBD5E1'
    COLOR_ACCENT    = '3B82F6'

    thin_border = Border(
        left=Side(style='thin', color=COLOR_BORDER),
        right=Side(style='thin', color=COLOR_BORDER),
        top=Side(style='thin', color=COLOR_BORDER),
        bottom=Side(style='thin', color=COLOR_BORDER),
    )
    accent_bottom_border = Border(
        left=Side(style='thin', color=COLOR_BORDER),
        right=Side(style='thin', color=COLOR_BORDER),
        top=Side(style='thin', color=COLOR_BORDER),
        bottom=Side(style='medium', color=COLOR_ACCENT),
    )

    ws.merge_cells('A1:L1')
    title_cell = ws['A1']
    title_cell.value     = 'Clean Ticket Report'
    title_cell.font      = Font(name='Calibri', bold=True, size=14, color=COLOR_HEADER_BG)
    title_cell.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 24

    ws.merge_cells('A2:L2')
    info_parts = []
    if query:     info_parts.append(f"Search: {query}")
    if date_from: info_parts.append(f"From: {date_from}")
    if date_to:   info_parts.append(f"To: {date_to}")
    info_parts.append(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    info_parts.append(f"Total records: {qs.count()}")
    info_cell = ws['A2']
    info_cell.value     = '  |  '.join(info_parts)
    info_cell.font      = Font(name='Calibri', size=9, color='64748B')
    info_cell.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 6

    HEADERS = [
        ('No', 5), ('Ticket ID', 14), ('Parent', 14), ('Package', 14), ('Platform', 14),
        ('Process', 18), ('Stage', 14), ('Area', 16), ('Start Date', 13),
        ('Due Date', 13), ('CT (days)', 10), ('Status', 13),
    ]
    header_fill  = PatternFill('solid', fgColor=COLOR_HEADER_BG)
    header_font  = Font(name='Calibri', bold=True, size=9, color=COLOR_HEADER_FG)
    header_align = Alignment(horizontal='center', vertical='center')

    for col_idx, (label, width) in enumerate(HEADERS, start=1):
        cell = ws.cell(row=4, column=col_idx, value=label)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = header_align
        cell.border    = accent_bottom_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[4].height = 20

    center_align = Alignment(horizontal='center', vertical='center')
    left_align   = Alignment(horizontal='left',   vertical='center')
    ct_font      = Font(name='Calibri', size=9, bold=True, color='1D4ED8')
    base_font    = Font(name='Calibri', size=9)

    for i, t in enumerate(qs, start=1):
        row_num = i + 4
        item = {
            'ticket_key':         t.ticket_key,
            'parent_key':         t.parent_key,
            'platform':           t.platform,
            'package_name':       t.package_name,
            'predefined_process': t.predefined_process,
            'status':             t.status,
            'start_date':         t.start_date,
            'due_date':           t.due_date,
            'cycle_time':         t.cycle_time,
        }
        e      = enrich_ticket(item)
        ct_val = round(e['cycle_time'], 1) if e.get('cycle_time') else None
        row_fill = PatternFill('solid', fgColor=COLOR_ROW_ODD if i % 2 else COLOR_ROW_EVEN)

        values = [
            i, e.get('ticket_key',''), e.get('parent_key') or '',
            e.get('platform') or '', e.get('platform_group') or '',
            e.get('predefined_process') or '', e.get('process_stage') or '',
            e.get('process_area') or '', str(e.get('start_date') or ''),
            str(e.get('due_date') or ''), ct_val, e.get('status',''),
        ]
        aligns = [
            center_align, left_align, left_align, left_align, left_align,
            left_align, left_align, left_align,
            center_align, center_align, center_align, center_align,
        ]
        for col_idx, (val, aln) in enumerate(zip(values, aligns), start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.fill      = row_fill
            cell.border    = thin_border
            cell.alignment = aln
            cell.font      = ct_font if col_idx == 11 and val is not None else base_font
        ws.row_dimensions[row_num].height = 16

    ws.freeze_panes = 'A5'
    ws.auto_filter.ref = f"A4:L{ws.max_row}"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"report-{datetime.now().strftime('%Y%m%d-%H%M%S')}.xlsx"
    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ======================================================
# DASHBOARD STAFF
# ======================================================
@login_required
@role_required('staff')
def dashboard_staff(request):
    published     = get_published_syncs()
    total_ticket  = RawTicket.objects.filter(parent_key__isnull=False, sync_log__in=published).count()
    avg_ct        = RawTicket.objects.filter(cycle_time__isnull=False, sync_log__in=published).aggregate(avg=Avg('cycle_time'))['avg']
    avg_ct        = round(avg_ct, 1) if avg_ct else 0
    total_flagged = ErrorTicket.objects.count()

    context = {
        'username':      request.user.username,
        'role':          request.user.get_role_display(),
        'total_ticket':  total_ticket,
        'avg_ct':        avg_ct,
        'total_flagged': total_flagged,
        **get_dashboard_filter_options(),
    }
    return render(request, 'main/staff/dashboard_staff.html', context)


# ======================================================
# VIEW DATA STAFF
# ======================================================
@login_required
@role_required('staff')
def staff_view_data(request):
    query       = request.GET.get('q', '')
    page_number = request.GET.get('page', 1)
    per_page    = int(request.GET.get('per_page', 10))
    if per_page not in [10, 25, 50, 100]:
        per_page = 10

    if request.method == 'POST':
        action     = request.POST.get('action')
        ticket_key = request.POST.get('ticket_key')
        comment    = request.POST.get('comment', '')

        if action == 'flag':
            try:
                ticket = RawTicket.objects.get(ticket_key=ticket_key)
                ErrorTicket.objects.get_or_create(
                    ticket=ticket,
                    defaults={
                        'error_message': comment or 'Flagged by staff',
                        'flagged_by':    request.user,
                        'raw_payload':   {},
                    }
                )
                FlagHistory.objects.create(
                    ticket=ticket, flagged_by=request.user,
                    action='Flag', comment=comment,
                )
                messages.success(request, f"Ticket {ticket_key} has been flagged.")
            except RawTicket.DoesNotExist:
                messages.error(request, "Ticket not found.")
        return redirect('staff_view_data')

    last_sync      = SyncLog.objects.filter(status='success').order_by('-started_at').first()
    sync_data      = []
    paginator      = None
    page_obj       = None
    page_range     = []
    start_index    = 0
    end_index      = 0
    total_filtered = 0
    total_all_raw  = RawTicket.objects.filter(parent_key__isnull=False).count()

    if last_sync:
        #untuk membuat minus muncul di awal
        qs = RawTicket.objects.filter(
            parent_key__isnull=False,
            sync_log=last_sync,
        ).annotate(
            error_priority=Case(
                When(cycle_time__lt=0,        then=Value(0)),
                When(cycle_time__isnull=True, then=Value(1)),
                default=Value(2),
                output_field=IntegerField(),
            )
        ).order_by('error_priority', 'ticket_key')

        if query:
            qs = qs.filter(
                Q(ticket_key__icontains=query) |
                Q(predefined_process__icontains=query)
            )

        flagged_keys   = set(ErrorTicket.objects.values_list('ticket__ticket_key', flat=True))
        total_filtered = qs.count()

        paginator   = Paginator(qs, per_page)
        page_obj    = paginator.get_page(page_number)
        start_index = (page_obj.number - 1) * per_page + 1
        end_index   = min(page_obj.number * per_page, total_filtered)

        for t in page_obj.object_list:
            item = {
                'ticket_key':         t.ticket_key,
                'parent_key':         t.parent_key,
                'platform':           t.platform,
                'package_name':       t.package_name,
                'predefined_process': t.predefined_process,
                'status':             t.status,
                'start_date':         t.start_date,
                'due_date':           t.due_date,
                'cycle_time':         t.cycle_time,
                'has_ct':             t.cycle_time is not None,
                'is_flagged':         t.ticket_key in flagged_keys,
            }
            sync_data.append(enrich_ticket(item))
        page_range = get_page_range(page_obj.number, paginator.num_pages)

    # ── Data range string ─────────────────────────────────────
    BULAN_ID = {
        1:'Januari', 2:'Februari', 3:'Maret',    4:'April',
        5:'Mei',     6:'Juni',     7:'Juli',      8:'Agustus',
        9:'September',10:'Oktober',11:'November', 12:'Desember',
    }
    dr = RawTicket.objects.filter(start_date__isnull=False).aggregate(
        earliest=Min('start_date'), latest=Max('start_date')
    )
    e, l = dr['earliest'], dr['latest']
    data_range_str = (
        f"{e.day} {BULAN_ID[e.month]} {e.year} — {l.day} {BULAN_ID[l.month]} {l.year}"
        if e and l else None
    )

    total_all_sync  = RawTicket.objects.filter(
        parent_key__isnull=False, sync_log=last_sync
    ).count() if last_sync else 0
    flagged_count   = ErrorTicket.objects.count()
    clean_count     = total_all_sync - flagged_count
    context = {
        'sync_data':      sync_data,
        'query':          query,
        'per_page':       per_page,
        'page_obj':       page_obj,
        'paginator':      paginator,
        'page_range':     page_range,
        'start_index':    start_index,
        'end_index':      end_index,
        'total_filtered': total_filtered,
        'total_all':      total_all_raw,
        'last_sync':      last_sync,
        'data_range_str': data_range_str,
        'clean_count':    clean_count,
        'flagged_count':  flagged_count,
    }
    return render(request, 'main/staff/view_data.html', context)


# ======================================================
# REPORTS STAFF
# ======================================================
@login_required
@role_required('staff')
def staff_reports(request):
    query       = request.GET.get('q', '')
    date_from   = request.GET.get('date_from', '')
    date_to     = request.GET.get('date_to', '')
    page_number = request.GET.get('page', 1)
    per_page    = int(request.GET.get('per_page', 10))
    if per_page not in [10, 25, 50, 100]:
        per_page = 10

    published_syncs = SyncLog.objects.filter(status='published').values_list('id', flat=True)
    qs = RawTicket.objects.filter(
        sync_log__in=published_syncs, parent_key__isnull=False
    ).order_by('ticket_key')

    if query:
        qs = qs.filter(
            Q(ticket_key__icontains=query) |
            Q(platform__icontains=query)   |
            Q(predefined_process__icontains=query)
        )
    if date_from:
        qs = qs.filter(start_date__gte=date_from)
    if date_to:
        qs = qs.filter(start_date__lte=date_to)

    seen, unique_ids = set(), []
    for t in qs.values('id', 'ticket_key'):
        if t['ticket_key'] not in seen:
            seen.add(t['ticket_key'])
            unique_ids.append(t['id'])
    qs = RawTicket.objects.filter(id__in=unique_ids).order_by('ticket_key')

    total_main     = RawTicket.objects.filter(sync_log__in=published_syncs, parent_key__isnull=True).count()
    total_sub      = RawTicket.objects.filter(sync_log__in=published_syncs, parent_key__isnull=False).count()
    total_filtered = qs.count()

    paginator   = Paginator(qs, per_page)
    page_obj    = paginator.get_page(page_number)
    start_index = (page_obj.number - 1) * per_page + 1
    end_index   = min(page_obj.number * per_page, total_filtered)

    data = []
    for t in page_obj.object_list:
        item = {
            'ticket_key':         t.ticket_key,
            'parent_key':         t.parent_key,
            'platform':           t.platform,
            'package_name':       t.package_name,
            'predefined_process': t.predefined_process,
            'status':             t.status,
            'start_date':         t.start_date,
            'due_date':           t.due_date,
            'cycle_time':         t.cycle_time,
            'has_ct':             t.cycle_time is not None,
        }
        data.append(enrich_ticket(item))

    context = {
        'data':           data,
        'query':          query,
        'date_from':      date_from,
        'date_to':        date_to,
        'per_page':       per_page,
        'page_obj':       page_obj,
        'paginator':      paginator,
        'page_range':     get_page_range(page_obj.number, paginator.num_pages),
        'start_index':    start_index,
        'end_index':      end_index,
        'total_filtered': total_filtered,
        'total_main':     total_main,
        'total_sub':      total_sub,
    }
    return render(request, 'main/staff/reports.html', context)


# ======================================================
# DASHBOARD MANAGEMENT
# ======================================================
@login_required
@role_required('management')
def dashboard_management(request):
    published    = get_published_syncs()
    total_ticket = RawTicket.objects.filter(parent_key__isnull=False, sync_log__in=published).count()
    avg_ct       = RawTicket.objects.filter(cycle_time__isnull=False, sync_log__in=published).aggregate(avg=Avg('cycle_time'))['avg']
    avg_ct       = round(avg_ct, 1) if avg_ct else 0

    context = {
        'username':     request.user.username,
        'role':         request.user.get_role_display(),
        'total_ticket': total_ticket,
        'avg_ct':       avg_ct,
        **get_dashboard_filter_options(),
    }
    return render(request, 'main/management/dashboard_management.html', context)


# ======================================================
# REPORTS MANAGEMENT
# ======================================================
@login_required
@role_required('management')
def management_reports(request):
    query       = request.GET.get('q', '')
    date_from   = request.GET.get('date_from', '')
    date_to     = request.GET.get('date_to', '')
    page_number = request.GET.get('page', 1)
    per_page    = int(request.GET.get('per_page', 10))
    if per_page not in [10, 25, 50, 100]:
        per_page = 10

    published_syncs = SyncLog.objects.filter(status='published').values_list('id', flat=True)
    qs = RawTicket.objects.filter(
        sync_log__in=published_syncs, parent_key__isnull=False
    ).order_by('ticket_key')

    if query:
        qs = qs.filter(
            Q(ticket_key__icontains=query) |
            Q(platform__icontains=query)   |
            Q(predefined_process__icontains=query)
        )
    if date_from:
        qs = qs.filter(start_date__gte=date_from)
    if date_to:
        qs = qs.filter(start_date__lte=date_to)

    seen, unique_ids = set(), []
    for t in qs.values('id', 'ticket_key'):
        if t['ticket_key'] not in seen:
            seen.add(t['ticket_key'])
            unique_ids.append(t['id'])
    qs = RawTicket.objects.filter(id__in=unique_ids).order_by('ticket_key')

    total_main     = RawTicket.objects.filter(sync_log__in=published_syncs, parent_key__isnull=True).count()
    total_sub      = RawTicket.objects.filter(sync_log__in=published_syncs, parent_key__isnull=False).count()
    total_filtered = qs.count()

    paginator   = Paginator(qs, per_page)
    page_obj    = paginator.get_page(page_number)
    start_index = (page_obj.number - 1) * per_page + 1
    end_index   = min(page_obj.number * per_page, total_filtered)

    data = []
    for t in page_obj.object_list:
        item = {
            'ticket_key':         t.ticket_key,
            'parent_key':         t.parent_key,
            'platform':           t.platform,
            'package_name':       t.package_name,
            'predefined_process': t.predefined_process,
            'status':             t.status,
            'start_date':         t.start_date,
            'due_date':           t.due_date,
            'cycle_time':         t.cycle_time,
            'has_ct':             t.cycle_time is not None,
        }
        data.append(enrich_ticket(item))

    context = {
        'data':           data,
        'query':          query,
        'date_from':      date_from,
        'date_to':        date_to,
        'per_page':       per_page,
        'page_obj':       page_obj,
        'paginator':      paginator,
        'page_range':     get_page_range(page_obj.number, paginator.num_pages),
        'start_index':    start_index,
        'end_index':      end_index,
        'total_filtered': total_filtered,
        'total_main':     total_main,
        'total_sub':      total_sub,
    }
    return render(request, 'main/management/reports.html', context)


# ======================================================
# CHART API
# ======================================================
from django.db.models import Avg, Count, F

@login_required
def chart_data_api(request):
    from main.process_groups import get_process_info
    published = get_published_syncs()

    qs = RawTicket.objects.filter(
        parent_key__isnull=False,
        cycle_time__isnull=False,
        platform__isnull=False,
        sync_log__in=published,
    ).exclude(platform='')

    ct_per_platform = (
        qs.values('platform')
          .annotate(avg_ct=Avg('cycle_time'), count=Count('id'))
          .order_by('platform')
    )
    chart1 = {'labels': [], 'data': [], 'counts': []}
    for row in ct_per_platform:
        chart1['labels'].append(row['platform'])
        chart1['data'].append(round(row['avg_ct'], 2))
        chart1['counts'].append(row['count'])

    from django.db.models.functions import ExtractYear
    ct_per_platform_year = (
        qs.annotate(year=ExtractYear('start_date'))
          .filter(year__isnull=False)
          .values('platform', 'year')
          .annotate(avg_ct=Avg('cycle_time'))
          .order_by('platform', 'year')
    )
    platform_year_map = {}
    years_set = set()
    for row in ct_per_platform_year:
        p = row['platform']; y = str(row['year'])
        years_set.add(y)
        if p not in platform_year_map: platform_year_map[p] = {}
        platform_year_map[p][y] = round(row['avg_ct'], 2)

    years_sorted = sorted(years_set)
    platforms    = sorted(platform_year_map.keys())
    COLORS = [
        '#2563eb','#16a34a','#dc2626','#d97706','#7c3aed','#0891b2',
        '#be185d','#059669','#ea580c','#4338ca','#0d9488','#b45309','#6d28d9',
    ]
    chart2_datasets = []
    for i, platform in enumerate(platforms):
        chart2_datasets.append({
            'label':           platform,
            'data':            [platform_year_map[platform].get(y) for y in years_sorted],
            'borderColor':     COLORS[i % len(COLORS)],
            'backgroundColor': COLORS[i % len(COLORS)] + '33',
            'tension':         0.3,
            'fill':            False,
        })
    chart2 = {'labels': years_sorted, 'datasets': chart2_datasets}

    ct_per_process_platform = (
        qs.filter(predefined_process__isnull=False)
          .exclude(predefined_process='')
          .values('platform', 'predefined_process')
          .annotate(avg_ct=Avg('cycle_time'))
          .order_by('platform', 'predefined_process')
    )
    process_platform_map = {}
    for row in ct_per_process_platform:
        p = row['platform']; pr = row['predefined_process']
        if p not in process_platform_map: process_platform_map[p] = {}
        process_platform_map[p][pr] = round(row['avg_ct'], 2)

    chart3 = {'platforms': sorted(process_platform_map.keys()), 'data': process_platform_map}

    estimasi_data = {}
    for row in ct_per_process_platform:
        p = row['platform']; pr = row['predefined_process']
        if p not in estimasi_data: estimasi_data[p] = {}
        estimasi_data[p][pr] = round(row['avg_ct'], 2)

    all_processes = sorted(list(
        RawTicket.objects.filter(predefined_process__isnull=False)
        .exclude(predefined_process='')
        .values_list('predefined_process', flat=True).distinct()
    ))
    all_platforms = sorted(list(
        RawTicket.objects.filter(platform__isnull=False)
        .exclude(platform='')
        .values_list('platform', flat=True).distinct()
    ))
    chart4 = {'estimasi_data': estimasi_data, 'all_processes': all_processes, 'all_platforms': all_platforms}

    return JsonResponse({'chart1': chart1, 'chart2': chart2, 'chart3': chart3, 'chart4': chart4})


# ======================================================
# CYCLE TIME ANALYSIS DASHBOARD
# ======================================================
@login_required
def ct_analysis_dashboard(request):
    from main.process_groups import PROCESS_GROUP, PACKAGE_PLATFORM
    published = get_published_syncs()

    months_qs = RawTicket.objects.filter(
        parent_key__isnull=False,
        start_date__isnull=False,
        cycle_time__isnull=False,
        sync_log__in=published,
    ).dates('start_date', 'month')

    context = {
        'all_months':    sorted(set(f"{d.year}-{str(d.month).zfill(2)}" for d in months_qs)),
        'all_platforms': sorted(set(PACKAGE_PLATFORM.values())),
        'all_stages':    sorted(set(v[0] for v in PROCESS_GROUP.values())),
        'all_areas':     sorted(set(v[1] for v in PROCESS_GROUP.values())),
        'all_processes': sorted(set(PROCESS_GROUP.keys())),
    }
    return render(request, 'main/partials/ct_analysis.html', context)


@login_required
def ct_analysis_data(request):
    from main.process_groups import PROCESS_GROUP, PACKAGE_PLATFORM
    from collections import defaultdict
    published = get_published_syncs()

    f_year      = request.GET.get('year', '').strip()
    f_platforms = request.GET.getlist('platform')
    f_stages    = request.GET.getlist('stage')
    f_areas     = request.GET.getlist('area')
    f_processes = request.GET.getlist('process')

    qs = RawTicket.objects.filter(
        parent_key__isnull=False,
        cycle_time__isnull=False,
        start_date__isnull=False,
        sync_log__in=published,
    ).exclude(predefined_process__isnull=True).exclude(predefined_process='') \
     .exclude(platform__isnull=True).exclude(platform='')

    force_monthly = False
    if f_year:
        try:
            qs = qs.filter(start_date__year=int(f_year))
            force_monthly = True
        except ValueError:
            pass
    if f_processes:
        qs = qs.filter(predefined_process__in=f_processes)

    rows          = qs.values('platform', 'predefined_process', 'cycle_time', 'start_date')
    _VALID_GROUPS = set(PACKAGE_PLATFORM.values())
    enriched = []
    for r in rows:
        proc        = r['predefined_process']
        plat        = r['platform']
        stage, area = PROCESS_GROUP.get(proc, ('Other', 'Other'))
        pg          = plat if plat in _VALID_GROUPS else 'Other'
        if f_platforms and pg    not in f_platforms: continue
        if f_stages    and stage not in f_stages:    continue
        if f_areas     and area  not in f_areas:     continue
        sd = r['start_date']
        enriched.append({
            'pg': pg, 'proc': proc, 'stage': stage, 'area': area,
            'ct': float(r['cycle_time']),
            'year':  str(sd.year),
            'month': f"{sd.year}-{str(sd.month).zfill(2)}",
        })

    all_years_db = sorted(set(
        str(d.year) for d in RawTicket.objects.filter(
            parent_key__isnull=False, start_date__isnull=False,
        ).dates('start_date', 'year')
    ))

    if not enriched:
        return JsonResponse({
            'col_keys': [], 'col_mode': {}, 'months': [],
            'platform_groups': [], 'all_years': all_years_db,
            'all_areas': [], 'all_processes': [],
            'pivot1': {}, 'pivot2': {},
            'chart': {'labels': [], 'datasets': []},
        })

    if force_monthly:
        col_keys = sorted(set(r['month'] for r in enriched))
        col_mode = {ck: 'monthly' for ck in col_keys}
    else:
        col_keys = sorted(set(r['year'] for r in enriched))
        col_mode = {ck: 'yearly' for ck in col_keys}

    def gcol(r):
        return r['year'] if col_mode.get(r['year']) == 'yearly' else r['month']

    p1_raw = defaultdict(lambda: defaultdict(list))
    for r in enriched:
        p1_raw[(r['pg'], r['stage'], r['area'], r['proc'])][gcol(r)].append(r['ct'])

    pivot1 = {}
    for (pg, stage, area, proc), col_data in p1_raw.items():
        pivot1.setdefault(pg, {}).setdefault(stage, {}).setdefault(area, {})
        apc, av = {}, []
        for ck in col_keys:
            vals = col_data.get(ck, [])
            if vals:
                apc[ck] = round(sum(vals)/len(vals), 1); av.extend(vals)
        apc['grand_total'] = round(sum(av)/len(av), 1) if av else None
        pivot1[pg][stage][area][proc] = apc

    pgs_in_data = sorted(set(r['pg'] for r in enriched))
    p2_raw = defaultdict(lambda: defaultdict(list))
    for r in enriched:
        p2_raw[(r['area'], r['proc'])][r['pg']].append(r['ct'])

    pivot2 = {}
    for (area, proc), pg_data in p2_raw.items():
        pivot2.setdefault(area, {})
        pga, av = {}, []
        for pg in pgs_in_data:
            vals = pg_data.get(pg, [])
            if vals:
                pga[pg] = round(sum(vals)/len(vals), 1); av.extend(vals)
        pga['grand_total'] = round(sum(av)/len(av), 1) if av else None
        pivot2[area][proc] = pga

    AREA_COLORS = {
        'Die Attach':'#2563eb','Wire Bond':'#16a34a','Molding':'#dc2626',
        'Trim & Form':'#d97706','Marking':'#7c3aed','Plating':'#0891b2',
        'Ball Mount':'#be185d','Quality Control':'#059669','Testing':'#ea580c',
        'Reliability Test':'#4338ca','Packing':'#0d9488','Shipment':'#b45309',
        'Wafer Preparation':'#6d28d9','Other':'#9ca3af',
    }
    chart_raw = defaultdict(lambda: defaultdict(list))
    for r in enriched:
        chart_raw[r['area']][gcol(r)].append(r['ct'])

    datasets = []
    for area in sorted(chart_raw.keys()):
        pts = [
            round(sum(chart_raw[area].get(ck,[]))/max(len(chart_raw[area].get(ck,[])),1), 1)
            if chart_raw[area].get(ck) else 0
            for ck in col_keys
        ]
        nz = [x for x in pts if x]
        pts.append(round(sum(nz)/len(nz), 1) if nz else 0)
        datasets.append({
            'label': area, 'data': pts,
            'backgroundColor': AREA_COLORS.get(area, '#9ca3af') + 'cc',
            'borderColor':     AREA_COLORS.get(area, '#9ca3af'),
            'borderWidth': 1,
        })

    return JsonResponse({
        'col_keys': col_keys, 'col_mode': col_mode, 'months': col_keys,
        'use_yearly': not force_monthly,
        'platform_groups': pgs_in_data, 'all_years': all_years_db,
        'all_areas':    sorted(pivot2.keys()),
        'all_processes': sorted(set(p for ad in pivot2.values() for p in ad.keys())),
        'pivot1': pivot1, 'pivot2': pivot2,
        'chart': {'labels': col_keys + ['Grand Total'], 'datasets': datasets},
    })